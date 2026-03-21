"""Excel export — Phase 6.

Generates an .xlsx file matching the Vaktlista_Geiranger_Sjokolade spreadsheet format.
The month is split into two halves (days 1–15, days 16–end) stacked vertically.
"""
from __future__ import annotations

import calendar
import io
from datetime import date

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.demand.forecaster import DailyDemand
from src.models.employee import EmployeeRead
from src.models.schedule import AssignmentRead, ScheduleRead
from src.models.shift_template import ShiftTemplateRead

# ── Color palette (hex without #) ────────────────────────────────────────────
COLORS = {
    "day_off":           "FFB3B3",  # red
    "part_time_avail":   "FFD699",  # orange
    "working":           "B3FFB3",  # light green
    "header_day":        "D9E1F2",  # light blue
    "prod_header":       "F4B084",  # salmon — PRODUCTION section
    "cafe_header":       "9BC2E6",  # sky blue — CAFÉ section
    "cruise_header":     "C8E6FA",  # light blue — cruise section
    "cruise_row":        "E8F4FD",  # very light blue
    "title_bg":          "4A90D9",  # dark blue — title rows
    "legend_header":     "EEEEEE",  # grey
}

_THIN = Side(style="thin", color="AAAAAA")
_THIN_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_DAY_WIDTH = 5      # column width for day columns
_NAME_WIDTH = 18    # column width for employee name
_LEGEND_WIDTH = 28  # column width for shift legend


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _font(bold: bool = False, color: str = "000000", size: int = 10) -> Font:
    return Font(bold=bold, color=color, size=size)


def _align(h: str = "center", v: str = "center", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _apply(cell, value=None, fill=None, font=None, align=None, border=None):
    if value is not None:
        cell.value = value
    if fill is not None:
        cell.fill = fill
    if font is not None:
        cell.font = font
    if align is not None:
        cell.alignment = align
    if border is not None:
        cell.border = border


def _shift_hours(shift: ShiftTemplateRead) -> float:
    s = shift.start_time.hour * 60 + shift.start_time.minute
    e = shift.end_time.hour * 60 + shift.end_time.minute
    return (e - s) / 60.0


# ── Main export function ─────────────────────────────────────────────────────

def export_schedule_to_excel(
    schedule: ScheduleRead,
    employees: list[EmployeeRead],
    demand: list[DailyDemand],
    shift_templates: list[ShiftTemplateRead],
) -> bytes:
    """Generate Excel workbook. Returns bytes ready for download."""
    month = schedule.month
    year = schedule.year
    month_name = calendar.month_name[month]
    _, days_in_month = calendar.monthrange(year, month)

    demand_map = {d.date: d for d in demand}
    assign_map: dict[tuple, str] = {
        (a.employee_id, a.date): a.shift_id for a in schedule.assignments
    }
    shift_map = {s.id: s for s in shift_templates}

    all_days = [date(year, month, d) for d in range(1, days_in_month + 1)]
    half1 = [d for d in all_days if d.day <= 15]
    half2 = [d for d in all_days if d.day > 15]

    prod_emps = sorted(
        [e for e in employees if e.role_capability in ("production", "both")],
        key=lambda e: (e.housing == "eidsdal", e.name),
    )
    cafe_emps = sorted(
        [e for e in employees if e.role_capability == "cafe"],
        key=lambda e: (e.housing == "eidsdal", e.name),
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{month_name} {year}"

    # Freeze top row
    ws.freeze_panes = "B1"

    # ── Title block ───────────────────────────────────────────────────────────
    ws.merge_cells("A1:P1")
    t = ws["A1"]
    _apply(t,
           value=f"Geiranger Sjokolade — {year}",
           fill=_fill(COLORS["title_bg"]),
           font=_font(bold=True, color="FFFFFF", size=14),
           align=_align("left"))

    ws.merge_cells("A2:P2")
    _apply(ws["A2"],
           value=f"Vaktplan {month_name} {year}",
           fill=_fill(COLORS["title_bg"]),
           font=_font(bold=True, color="FFFFFF", size=12),
           align=_align("left"))

    current_row = 4  # leave row 3 for opening hours (populated later)

    # ── Write each half ───────────────────────────────────────────────────────
    legend_written = False
    for half_idx, days in enumerate([half1, half2]):
        if not days:
            continue
        block_start = current_row
        current_row = _write_half(
            ws, current_row, days,
            prod_emps, cafe_emps,
            demand_map, assign_map, shift_map, shift_templates,
            write_legend=(not legend_written),
        )
        legend_written = True
        current_row += 2  # gap between halves

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Summary")
    _write_summary_sheet(ws_sum, schedule, employees, shift_templates)

    # ── Column widths on main sheet ───────────────────────────────────────────
    ws.column_dimensions["A"].width = _NAME_WIDTH
    for col in range(2, 18):
        ws.column_dimensions[get_column_letter(col)].width = _DAY_WIDTH
    ws.column_dimensions[get_column_letter(18)].width = _LEGEND_WIDTH

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Half-block writer ─────────────────────────────────────────────────────────

def _write_half(
    ws,
    start_row: int,
    days: list[date],
    prod_emps: list[EmployeeRead],
    cafe_emps: list[EmployeeRead],
    demand_map: dict,
    assign_map: dict,
    shift_map: dict,
    shift_templates: list[ShiftTemplateRead],
    write_legend: bool,
) -> int:
    """Write one half of the month. Returns the last row written."""
    row = start_row
    num_day_cols = len(days)
    legend_col = 2 + num_day_cols + 1  # one gap column

    # ── Day header rows ───────────────────────────────────────────────────────
    # Row: day-of-week abbreviations
    ws.cell(row=row, column=1, value="").fill = _fill(COLORS["header_day"])
    for i, d in enumerate(days, start=2):
        cell = ws.cell(row=row, column=i)
        is_weekend = d.weekday() >= 5
        _apply(cell,
               value=d.strftime("%a"),
               fill=_fill("C0D0F0" if is_weekend else COLORS["header_day"]),
               font=_font(bold=True, size=9),
               align=_align(),
               border=_THIN_BORDER)
    row += 1

    # Row: day numbers
    ws.cell(row=row, column=1, value="").fill = _fill(COLORS["header_day"])
    for i, d in enumerate(days, start=2):
        cell = ws.cell(row=row, column=i)
        is_weekend = d.weekday() >= 5
        _apply(cell,
               value=d.day,
               fill=_fill("C0D0F0" if is_weekend else COLORS["header_day"]),
               font=_font(bold=True, size=9),
               align=_align(),
               border=_THIN_BORDER)
    row += 1

    # ── PRODUCTION section ────────────────────────────────────────────────────
    row = _write_section_header(ws, row, "PRODUCTION", num_day_cols, COLORS["prod_header"])
    for emp in prod_emps:
        row = _write_employee_row(ws, row, emp, days, assign_map, shift_map)

    # ── CAFÉ section ──────────────────────────────────────────────────────────
    row = _write_section_header(ws, row, "CAFÉ", num_day_cols, COLORS["cafe_header"])
    for emp in cafe_emps:
        row = _write_employee_row(ws, row, emp, days, assign_map, shift_map)

    # ── Cruise info section ───────────────────────────────────────────────────
    row += 1
    row = _write_cruise_rows(ws, row, days, demand_map, num_day_cols)

    # ── Shift legend (right side, starting from half block top) ──────────────
    if write_legend:
        _write_shift_legend(ws, start_row, legend_col, shift_templates)

    return row


def _write_section_header(ws, row: int, label: str, num_day_cols: int, color: str) -> int:
    cell = ws.cell(row=row, column=1, value=label)
    _apply(cell,
           fill=_fill(color),
           font=_font(bold=True),
           align=_align("left"),
           border=_THIN_BORDER)
    for col in range(2, 2 + num_day_cols):
        ws.cell(row=row, column=col).fill = _fill(color)
        ws.cell(row=row, column=col).border = _THIN_BORDER
    return row + 1


def _write_employee_row(
    ws,
    row: int,
    emp: EmployeeRead,
    days: list[date],
    assign_map: dict,
    shift_map: dict,
) -> int:
    is_eidsdal = emp.housing == "eidsdal"
    is_pt = emp.employment_type == "part_time"
    name = emp.name + (" 🏔" if is_eidsdal else "") + (" 🚗" if emp.driving_licence else "")

    name_cell = ws.cell(row=row, column=1, value=name)
    name_bg = "FFF3E0" if is_eidsdal else "F8F8F8"
    _apply(name_cell,
           fill=_fill(name_bg),
           font=_font(bold=True, size=9),
           align=_align("left"),
           border=_THIN_BORDER)

    for i, d in enumerate(days, start=2):
        cell = ws.cell(row=row, column=i)
        is_avail = emp.availability_start <= d <= emp.availability_end

        if not is_avail:
            _apply(cell, fill=_fill("FFFFFF"), border=_THIN_BORDER)
            continue

        shift_id = assign_map.get((emp.id, d))
        if shift_id is None or shift_id == "off":
            # Day off
            color = COLORS["part_time_avail"] if is_pt else COLORS["day_off"]
            _apply(cell, value="", fill=_fill(color), border=_THIN_BORDER, align=_align())
        else:
            _apply(cell,
                   value=shift_id,
                   fill=_fill(COLORS["working"]),
                   font=_font(bold=True, size=9),
                   align=_align(),
                   border=_THIN_BORDER)

    return row + 1


def _write_cruise_rows(ws, row: int, days: list[date], demand_map: dict, num_day_cols: int) -> int:
    labels = ["Ships", "Port", "Count", "Arrival"]
    fields = ["ships", "port", "count", "arrival"]

    # Build data per day
    day_data: dict[date, dict] = {}
    for d in days:
        dd = demand_map.get(d)
        if dd and dd.ships_today:
            day_data[d] = {
                "ships": "\n".join(s.ship_name[:16] for s in dd.ships_today),
                "port": "\n".join(str(s.port).split("_")[0] for s in dd.ships_today),
                "count": str(len(dd.ships_today)),
                "arrival": "\n".join(str(s.arrival_time)[:5] for s in dd.ships_today),
            }

    # Header
    cell = ws.cell(row=row, column=1, value="CRUISE INFO")
    _apply(cell,
           fill=_fill(COLORS["cruise_header"]),
           font=_font(bold=True),
           align=_align("left"),
           border=_THIN_BORDER)
    for col in range(2, 2 + num_day_cols):
        ws.cell(row=row, column=col).fill = _fill(COLORS["cruise_header"])
        ws.cell(row=row, column=col).border = _THIN_BORDER
    row += 1

    for lbl, field in zip(labels, fields):
        lbl_cell = ws.cell(row=row, column=1, value=lbl)
        _apply(lbl_cell,
               fill=_fill(COLORS["cruise_header"]),
               font=_font(bold=True, size=9),
               align=_align("left"),
               border=_THIN_BORDER)
        for i, d in enumerate(days, start=2):
            cell = ws.cell(row=row, column=i)
            data = day_data.get(d, {})
            value = data.get(field, "")
            _apply(cell,
                   value=value,
                   fill=_fill(COLORS["cruise_row"]) if value else _fill("FFFFFF"),
                   font=_font(size=8),
                   align=_align(wrap=True),
                   border=_THIN_BORDER)
            if value and "\n" in value:
                ws.row_dimensions[row].height = 28
        row += 1

    return row


def _write_shift_legend(ws, start_row: int, col: int, shifts: list[ShiftTemplateRead]) -> None:
    header = ws.cell(row=start_row, column=col, value="SHIFT LEGEND")
    _apply(header,
           fill=_fill(COLORS["title_bg"]),
           font=_font(bold=True, color="FFFFFF"),
           align=_align("left"),
           border=_THIN_BORDER)

    row = start_row + 1
    cafe_shifts = [s for s in shifts if s.role == "cafe"]
    prod_shifts = [s for s in shifts if s.role == "production"]

    for section_label, section_shifts, color in [
        ("CAFÉ", cafe_shifts, COLORS["cafe_header"]),
        ("PRODUCTION", prod_shifts, COLORS["prod_header"]),
    ]:
        # Section sub-header
        sh = ws.cell(row=row, column=col, value=section_label)
        _apply(sh, fill=_fill(color), font=_font(bold=True, size=9),
               align=_align("left"), border=_THIN_BORDER)
        row += 1

        for s in sorted(section_shifts, key=lambda x: x.id):
            start = str(s.start_time)[:5]
            end = str(s.end_time)[:5]
            hours = _shift_hours(s)
            text = f"{s.id:>3} — {s.label}  {start}–{end}  ({hours:.1f}h)"
            cell = ws.cell(row=row, column=col, value=text)
            _apply(cell,
                   fill=_fill("FAFAFA"),
                   font=_font(size=9),
                   align=_align("left"),
                   border=_THIN_BORDER)
            row += 1

    ws.column_dimensions[get_column_letter(col)].width = _LEGEND_WIDTH


# ── Summary sheet ─────────────────────────────────────────────────────────────

def _write_summary_sheet(
    ws,
    schedule: ScheduleRead,
    employees: list[EmployeeRead],
    shift_templates: list[ShiftTemplateRead],
) -> None:
    """Write per-employee stats to the Summary sheet."""
    shift_h = {s.id: _shift_hours(s) for s in shift_templates}
    month_name = calendar.month_name[schedule.month]

    # Title
    ws.merge_cells("A1:H1")
    _apply(ws["A1"],
           value=f"Employee Summary — {month_name} {schedule.year}",
           fill=_fill(COLORS["title_bg"]),
           font=_font(bold=True, color="FFFFFF", size=12),
           align=_align("left"))

    headers = ["Employee", "Role", "Type", "Days Worked", "Days Off",
               "Total Hours", "Contracted (est.)", "Overtime"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=h)
        _apply(cell, fill=_fill(COLORS["header_day"]),
               font=_font(bold=True, size=9), align=_align(), border=_THIN_BORDER)

    assign_map: dict[tuple, str] = {
        (a.employee_id, a.date): a.shift_id for a in schedule.assignments
    }

    row = 3
    for emp in sorted(employees, key=lambda e: (e.role_capability, e.name)):
        total_h = sum(
            shift_h.get(sid, 0)
            for (eid, _), sid in assign_map.items()
            if eid == emp.id and sid != "off"
        )
        days_worked = sum(
            1 for (eid, _), sid in assign_map.items()
            if eid == emp.id and sid != "off"
        )
        days_off = sum(
            1 for (eid, _), sid in assign_map.items()
            if eid == emp.id and sid == "off"
        )
        contracted = emp.contracted_hours * 4.33
        overtime = max(0.0, total_h - contracted)

        values = [
            emp.name, emp.role_capability, emp.employment_type,
            days_worked, days_off,
            round(total_h, 1), round(contracted, 0), round(overtime, 1),
        ]
        ot_bg = "FFB3B3" if overtime > 0 else "FFFFFF"
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=val)
            _apply(cell,
                   fill=_fill(ot_bg),
                   font=_font(size=9, bold=(overtime > 0)),
                   align=_align(),
                   border=_THIN_BORDER)
        row += 1

    # Column widths
    for col, width in zip("ABCDEFGH", [18, 12, 12, 10, 10, 12, 14, 10]):
        ws.column_dimensions[col].width = width
